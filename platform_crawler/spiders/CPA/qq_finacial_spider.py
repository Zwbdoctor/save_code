"""
cpa qq 财务爬虫 ---- http://e.qq.com
"""
from re import search
import logging
import os
import time

from platform_crawler.spiders.pylib.login_qq_common import LoginQQ
from platform_crawler.spiders.pylib.task_process import TaskProcess
# from platform_crawler.spiders.pylib.base_crawler import BaseCrawler
from platform_crawler.utils.utils import Util
# from platform_crawler.spiders.pylib.get_pwd import get_pwd


logger = None
crawler = None
base_header = {
    'accept': "application/json, text/javascript, */*; q=0.01",
    'cookie': None,
    'Referer': "https://e.qq.com/atlas/5713092/account/list",
    'user-agent': "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3569.0 Safari/537.36",
}


class QQFinancialSpider(TaskProcess):

    def __init__(self, user_info):
        global logger, crawler
        log_name = 'YYBHLCPD'
        logger = logging.getLogger(log_name + '.qq')
        self.base_params = {"mod": 'aduser', 'act': 'info', 'unicode': 'true', 'g_tk': None, 'owner': None}
        self.uid = None
        self.cookies = {}
        self.gtk = None
        self.acc = user_info.get('account')
        self.platform = user_info.get('platform')
        super().__init__(spider=log_name)

    def getGTK(self, skey):
        hashes = 5381
        for letter in skey:
            hashes += (hashes << 5) + ord(letter)
        # return hashes & 0x7fffffff
        return hashes & 2147483647

    def deal_result(self, result, json_str=False):
        res = self.base_result(json_str=json_str)
        if not res.get('succ'):
            return res
        if res.get('msg').get('ret') != 0:
            logger.error(res.get('msg').get('msg'))
            return False
        else:
            return res.get('msg').get('data')

    def get_user_detail(self):
        url = 'https://e.qq.com/ec/api.php'
        skey = self.cookies.get('gdt_protect') if self.cookies.get('gdt_protect') else self.cookies.get('skey')
        self.cookies = '; '.join(['%s=%s' % (k, v) for k, v in self.cookies.items()])
        self.gtk = str(self.getGTK(skey))
        self.base_params.update({"g_tk": self.gtk, "owner": self.uid})
        base_header['cookie'] = self.cookies
        res = self.deal_result(self.execute('GET', url, params=self.base_params, verify=False, headers=base_header,
                                            print_payload=False), json_str=True)
        if not res:
            return False
        logger.info('---- company_name: %s' % res.get('uname'))
        self.company_name = res.get('uname')
        return True

    def get_account_type(self):
        url = "https://e.qq.com/ec/api.php"
        querystring = {"mod": "account", "act": "dashboard", "owner": self.uid, "unicode": "true", "g_tk": self.gtk}
        res = self.deal_result(self.execute('GET', url, params=querystring, verify=False, print_payload=False), json_str=True)
        if not res:
            return res
        app_ids = [{'app_id': e.get('app_id'), 'account_name': e.get('account_name')} for e in res.get('accounts')]
        logger.info('account_type----%s' % app_ids)
        return app_ids

    def get_data_process(self):
        app_ids = self.get_account_type()
        content = []
        mths, dates = Util().make_dates(ys=2016, ms=1, ye=2017, me=12)
        for sd, ed in dates:
            logger.info('date range ---- %s~%s' % (sd, ed))
            for i in app_ids:
                content.extend(self.get_data(i, sd, ed))
        return content

    def get_data(self, account_type, sd, ed, content=None, page=1):
        url = "https://e.qq.com/ec/api.php"
        querystring = {"mod": "account", 'act': 'costlist', "format": "json", "page": page, "pagesize": "20",
                       "sdate": sd, "edate": ed, 'accounttype': account_type.get('app_id')}
        self.base_params.update(querystring)
        res = self.deal_result(self.execute('GET', url, params=self.base_params, verify=False, print_payload=False), json_str=True)
        if res:
            if len(res.get('list')) == 0:
                return []
            data = [] if not content else content
            for e in res.get('list'):
                e.update({'user_type': account_type})
                data.append(e)
            # data.extend([e.update({'user_type': account_type}) for e in res.get('list')])
            if res.get('conf').get('next_page'):
                page += 1
                time.sleep(0.25)
                logger.info('got page %s  --- next_page %s' % (page, res.get('conf').get('next_page')))
                return self.get_data(account_type, sd, ed, content=data, page=page)
            return data
        return res

    def save(self, data):
        from openpyxl import load_workbook
        xlspath = os.path.join(os.path.abspath('./save_data/xlsx_files'), '%s.xlsx' % self.platform)
        workbook = load_workbook(xlspath)
        datas = [[self.company_name, self.acc, i.get('date'), i.get('balance'), i.get('amount'), i.get('desc'),
                  i.get('user_type').get('account_name')] for i in data]
        heading = ['公司名', '账号', '日期', '存入（元）', '支出（元）', '备注', '账户类型']
        rows = 1
        try:
            table = workbook.create_sheet(self.company_name)
            datas.insert(0, heading)
        except:
            table = workbook.get_sheet_by_name(self.company_name)
            rows = len(table.rows)
        for rowidx,row in enumerate(datas, rows):
            for itemidx,item in enumerate(row, 1):
                table.cell(rowidx, itemidx, item)
        workbook.save(xlspath)
        logger.info('saved data complete..')

    def run_task(self, ui):
        # ui['password'] = get_pwd(ui.get('password'))
        lq = LoginQQ(ui).run_login()
        if not lq.get('succ'):
            return lq
        self.cookies = {e.get('name'): e.get('value') for e in lq.get('cookies')}
        self.uid = search(r'uid=(\d+)', self.cookies.get('dm_cookie')).group(1)
        if not self.get_user_detail():
            return {'succ': False}
        data = self.get_data_process()
        self.save(data)


        # self.upload_file(ui, cutime, dir_name)