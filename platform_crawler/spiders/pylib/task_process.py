import json
import os
import time

from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl import load_workbook, workbook
import selenium.webdriver.support.expected_conditions as EC

from platform_crawler.spiders.pylib.base_crawler import BaseCrawler
from platform_crawler.spiders.pylib.get_pwd import get_pwd
from platform_crawler.spiders.pylib.scp_client import init_dst_dir, upload_file
from platform_crawler.utils.utils import Util
from platform_crawler import settings

post_res_url = 'http://erp.btomorrow.cn/adminjson/ERP_ReportPythonCrawlerTask'
u = Util()
logger = None
try:
    with open(settings.join(settings.task_type_path), 'r', encoding='utf-8') as tp:
        task_type = tp.read().strip()
except Exception:
    task_type = ''
settings.GlobalVal.CUR_TASK_TYPE = task_type


class TaskProcess(BaseCrawler):

    def __init__(self, *args, is_cpa=False, user_info=None, get_img=True, **kwargs):
        self.dir_path = None
        self.acc = user_info.get('account')
        self.pwd = None
        self.user_info = user_info
        self.is_cpa = is_cpa
        self.platform = user_info.get('platform')
        self.login_obj = None
        self.d = None
        self.wait = None
        self.logger = None
        self.result_kwargs = {'has_data': 1, 'has_pic': 1}
        if self.platform:
            self.getLogger(self.platform)
        self.err_img_name = None
        self.dst_path = None
        self.is_get_img = get_img
        self.balance_data = None
        self._dates = None
        self.user_agent = None
        super().__init__(*args, spider=self.platform, **kwargs)

    def getLogger(self, platform):
        global logger
        log_path = settings.join(settings.LOGPATH, platform)
        if not os.path.exists(log_path):
            os.makedirs(log_path)
        settings.GlobalVal.CUR_MAIN_LOG_NAME = platform
        logger = u.record_log(log_path, platform)
        self.logger = logger
        return self.logger

    def init_browser(self, page_timeout=120, element_timeout=10):
        co = webdriver.ChromeOptions()
        # co.add_argument('--no-sandbox')
        co.add_argument('--disable-infobars')
        # co.add_argument('--disable-gpu')
        co.add_argument('--disable-background-networking')
        self.d = webdriver.Chrome(options=co)
        self.d.delete_all_cookies()
        self.d.set_page_load_timeout(page_timeout)
        self.d.set_script_timeout(page_timeout)
        self.d.maximize_window()
        self.user_agent = self.d.execute_script('return window.navigator.userAgent')
        self.wait = WebDriverWait(self.d, element_timeout)

    def wait_element(self, element_type, text, ec=EC.visibility_of_element_located, wait_time=10):
        if not self.wait:
            assert self.d is not None, 'WebDriverWait Object Need Initial WebDriver First'
            self.wait = WebDriverWait(self.d, wait_time)
        if wait_time != 10:
            self.wait._timeout = wait_time
        ele = self.wait.until(ec((element_type, text)))
        return ele

    def upload_file(self, ui=None):
        if ui:  # 当self.run 抛出异常，需要重新初始化路径数据
            self.init_paths(ui)
        # 上传
        if not upload_file(self.dir_path, self.platform, is_cpa=self.is_cpa):
            return {'succ': False, 'msg': 'upload failed'}
        return {'succ': True}

    def login_part(self, *args, **kwargs):
        ...

    def deal_login_result(self, login_res):
        ...

    @property
    def get_dates(self):
        if not self._dates:
            dates = self.user_info.get('dates')
            ys, ms, ye, me = dates if dates else (None, None, None, None)
            mths, self._dates = u.make_dates(ms=ms, ys=ys, ye=ye, me=me)
            return self._dates
        return self._dates

    @get_dates.setter
    def get_dates(self, value):
        if isinstance(value, int):
            mths, self._dates = u.make_dates(dur=value)
        elif isinstance(value, tuple):
            ys, ms, ye, me = value
            mths, self._dates = u.make_dates(ms=ms, ys=ys, ye=ye, me=me, dur=value)
        else:
            raise Exception('value must be int for during or tuple for months list')

    def get_data_and_imgs(self, ui, **kwargs):
        # 图片与数据的获取控制
        res = self.get_data_part(ui, **kwargs)
        if isinstance(res, dict):
            if not res.get('succ', False):
                return res
        # 获取账户余额
        self.get_account_balance()
        if not self.is_get_img:
            return
        if self.result_kwargs.get('has_data') == 0:
            return
        return self.get_img_part(get_data_res=res, **kwargs)

    def get_data_part(self, *args, **kwargs):
        ...

    def get_img_part(self, *args, **kwargs):
        ...

    def login_and_get_data(self, ui):
        res = self.deal_login_result(self.login_part(ui))
        if res is not None:
            return res
        res = self.get_data_and_imgs(ui)
        if res is None:
            return {'succ': True}
        return res

    def get_balance(self):
        ...

    def get_account_balance(self):
        self.logger.info('There is no data since last month!')
        res = self.get_balance()
        if isinstance(res, dict) and not res.get('succ'):
            raise Exception(res)
        header, rows = self.parse_balance()
        if not rows:
            return
        # self.save_balance_to_xls(header, rows)
        self.save_balance_data(rows)

    def parse_balance(self, *args, **kwargs):
        return [], []

    def save_balance_data(self, balance):
        date = time.strftime('%Y-%m-%d')
        file_name = settings.join(settings.BAL_PATH, f'balance_data_{date}_{task_type}.json')
        try:
            with open(file_name, 'r') as reader:
                data = json.load(reader)
        except:
            data = {}
        if not data.get(self.platform):
            data[self.platform] = []
        if isinstance(balance, list):
            data.get(self.platform).extend(balance)
        else:
            data.get(self.platform).append({'account': self.acc, 'balance': balance})
        with open(file_name, 'w') as writer:
            json.dump(data, writer)

    def save_balance_to_xls(self, header: list, data: list):
        # open workbook
        date = time.strftime('%Y-%m-%d')
        dir_name = settings.join(settings.sd_path, 'balance_data')
        if not os.path.exists(dir_name):
            os.makedirs(dir_name)
        file_name = settings.join(dir_name, f'account_balance_%s.xlsx' % date)
        if os.path.exists(file_name):
            wb = load_workbook(filename=file_name)
        else:
            wb = workbook.Workbook()
        # open worksheet
        sheet_names = wb.sheetnames
        if self.platform in sheet_names:
            ws = wb[self.platform]
        else:
            ws = wb.create_sheet(title=self.platform)
            if 'Sheet' in sheet_names:
                del wb['Sheet']
        # write header
        if not ws['A1'].value:
            for col, i in enumerate(header, 1):
                ws.cell(row=1, column=col, value=i)
        # resort
        read_header = [x.value for x in next(ws.rows)]
        max_col = ws.max_column
        for c in header:
            if c not in read_header:
                max_col += 1
                ws.cell(1, max_col, value=c)
        for k, v in enumerate(data.copy()):
            data[k] = [v.get(x, 0) for x in read_header]
        # write
        for i in data:
            ws.append(i)
        # save
        wb.save(file_name)
        self.logger.info('Balance saved ok!')
        return

    def init_paths(self, ui):
        """初始化路径数据"""
        init_dst_dir(self.platform, is_cpa=self.is_cpa)
        # 创建文件夹
        cur_time = time.strftime('%Y-%m-%d')
        dir_name = '%(taskId)s_%(cTime)s_%(account)s' % {'taskId': ui['id'],
                                                         'cTime': time.strftime('%Y-%m-%d_%H-%M-%S'),
                                                         'account': ui.get('account')}
        self.dir_path = settings.join(settings.sd_path, self.platform, cur_time, dir_name)
        os.makedirs(self.dir_path)

        self.err_img_name = settings.join(self.dir_path, 'error_%s.jpg')
        self.dst_path = '/data/python/%s/%s/%s' % (self.platform, cur_time, dir_name)
        if self.is_cpa:
            self.dst_path = '/data/python/%s/%s/%s/%s' % ('CPA', self.platform, cur_time, dir_name)
        settings.DST_DIR = self.dst_path  # Add to login class for post res
        self.logger.info(f'Init local directory: {self.dir_path}')

    def run(self, ui):
        """for child class to rewrite"""
        # 初始化路径数据
        self.init_paths(ui)
        self.init_global_params()

        # 登陆 && 获取数据/图片
        try:
            res = self.login_and_get_data(ui)
            if res is not None and not res.get('succ'):  # 正常的报错场景
                settings.GlobalFunc.save_screen_shot(self.err_img_name % int(time.time()*1000))
        except Exception as er:
            settings.GlobalFunc.save_screen_shot(self.err_img_name % int(time.time() * 1000))  # 未知报错场景
            logger.error(er, exc_info=1)
            res = {'succ': False, 'msg': 'got unKnown error'}

        try:
            self.d.quit()
        except:
            pass

        # 检测是否有图片
        pics = [x for x in os.listdir(self.dir_path) if 'png' in x]
        if len(pics) == 0:
            self.result_kwargs['has_pic'] = 0
        # 上传结果
        self.upload_file()

        # 返回结果
        if res is None:
            return {'succ': True}
        if not res.get('succ'):
            return res
        return {'succ': True}

    def init_global_params(self):
        settings.GlobalVal.err_src_name = self.err_img_name

    def run_task(self, ui):
        ui['password'] = get_pwd(ui.get('password')).strip()
        self.pwd = ui.get('password')
        try:
            res = self.run(ui)
            # 上报服务器
            if res['succ']:
                # 成功
                if not self.post_res(ui['id'], ui['account'], status=3, **self.result_kwargs):
                    logger.error('----------after upload files, post result failed !!!!')
                    return False
                else:
                    logger.info('Upload success! Post result success!')
                    return True
            elif not res.get('succ') and res.get('invalid_account'):
                if not self.post_res(ui['id'], ui['account'], status=4, **self.result_kwargs):
                    return False
                return True
            else:
                logger.warning('爬虫逻辑错误，所有可能信息如下：')
                logger.error(res)
                if not self.post_res(ui['id'], ui['account'], status=5, **self.result_kwargs):
                    logger.error('---------- post error message failed !!!!')
                return False
                # logger.smtp('发送邮件')
        except Exception as e:
            logger.warning('Got an err about account %s, detail msg like this:' % ui['account'])
            logger.error(e, exc_info=1)
            settings.GlobalFunc.save_screen_shot(self.err_img_name % int(time.time()*1000))  # 账号无效场景
            self.upload_file(ui=ui)
            self.result_kwargs.update({'has_data': 0, 'has_pic': 0})
            self.post_res(ui['id'], ui['account'], status=5, **self.result_kwargs)
            return False

    def post_res(self, task_id, account, status=None, has_data=None, has_pic=None):
        """
        上报结果
        :param task_id:
        :param account:
        :param status: booltype, 是否成功
        :param has_data: int, 是否有数据
        :param has_pic: int, 是否有截图
        :return: 返回是否上报成功
        """
        post_data = {'taskId': task_id, 'errorCode': None, 'status': None, 'statusDesc': None, 'account': account,
                     'platform': self.platform, 'filePathCatalog': '', 'flag': task_type, 'isScreenshots': has_pic,
                     'isData': has_data}
        if status == 5:
            need_change = {'status': 5, 'statusDesc': '爬虫逻辑错误', 'filePathCatalog': self.dst_path}
        elif status == 3:
            need_change = {'status': 3, 'statusDesc': '成功'}
        else:
            # need_change = {'status': 5, 'statusDesc': '账号无效', 'filePathCatalog': self.dst_path,
            #                'isData': 0, 'isScreenshots': 0}
            need_change = {'status': 4, 'statusDesc': '账号无效', 'errorCode': 10000, 'filePathCatalog': self.dst_path,
                           'isData': 0, 'isScreenshots': 0}
        post_data.update(need_change)
        data = json.dumps(post_data)
        res = self.deal_result(
            self.execute('POST', post_res_url, data=data, headers={'Content-Type': 'application/json'}), json_str=True)
        if not res.get('succ'):
            # 上报失败
            return False
        else:
            # 上报成功
            logger.info('Post success! ret_msg: %s' % res.get('msg'))
            return True
