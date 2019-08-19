import os
from platform_crawler.settings import sd_path


def save(self, data):
    from openpyxl import load_workbook
    xlspath = os.path.join(sd_path, 'xlsx_files', '%s.xlsx' % self.platform)
    workbook = load_workbook(xlspath)
    datas = [[self.company_name, self.acc, i.get('date'), i.get('balance'), i.get('amount'), i.get('desc'),
              i.get('user_type').get('account_name')] for i in data]
    heading = ['公司名', '账号', '日期', '存入（元）', '支出（元）', '备注', '账户类型']
    rows = 1
    try:
        table = workbook.create_sheet(self.company_name)
        datas.insert(0, heading)
    except:
        table = workbook.get_sheet_by_name(self.company_name)
        rows = len(table.rows)
    for rowidx,row in enumerate(datas, rows):
        for itemidx,item in enumerate(row, 1):
            table.cell(rowidx, itemidx, item)
    workbook.save(xlspath)


print(os.path.dirname(__file__))
print(__file__)



